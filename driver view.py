import sys
import math
import numpy as np
import cv2
import pyzed.sl as sl
from datetime import datetime
from ultralytics import YOLO
from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import QApplication, QLabel, QVBoxLayout, QWidget
from PyQt5.QtGui import QPixmap
from PyQt5.QtCore import Qt, QTimer

model = YOLO("yolo-Weights/yolov8n.pt")

# Initialize Excel file
excel_file = "Downloads/distance_warning_log.xlsx"

# Check if the file exists, and create it if not
try:
    workbook = load_workbook(excel_file)
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Timestamp", "Distance (m)", "Warning Level", "Pre-Brake Percentage (%)"])
    workbook.save(excel_file)

def log_to_warning_file(warning):
    try:
        with open('current_warning.txt', 'w') as f:
            f.write(str(warning))
    except Exception as e:
        print(f"An error occurred while writing warning level to file: {e}")

def distance_level(distance):
    range_distance = [0, 1, 2, 3, 4]
    level_warning = [4, 3, 2, 1, 0]
    warning = 0
    for n in range_distance:
        if (distance >= n) and (distance < (n+1)):
            warning = level_warning[n]
        elif (distance < 0):
            warning = level_warning[0]
        elif (distance >= 5):
            warning = level_warning[4]
        else:
            pass
    return warning

def start_warning_sequence(warning):
    pre_brake_percentage = 0
    if warning == 0:
        print("Safe")
    elif warning == 1:
        print("Pre-Brake 30%")
        pre_brake_percentage = 30
    elif warning == 2:
        print("Pre-Brake 50%")
        pre_brake_percentage = 50
    elif warning == 3:
        print("Pre-Brake 70%")
        pre_brake_percentage = 70
    elif warning == 4:
        print("Brake Full")
        pre_brake_percentage = 100
    return pre_brake_percentage

def log_to_excel(distance, warning, pre_brake_percentage):
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        sheet.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), distance, warning, pre_brake_percentage])
        workbook.save(excel_file)
    except Exception as e:
        print(f"An error occurred while logging to Excel: {e}")

class BrakeLevelApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        # Set up the window
        self.setWindowTitle("Brake Level Indicator")
        self.setGeometry(100, 100, 400, 600)

        # Set up the image label
        self.label = QLabel(self)
        self.label.setAlignment(Qt.AlignCenter)
        self.updateImage(0)  # Initialize with the first image

        # Set up the layout
        layout = QVBoxLayout()
        layout.addWidget(self.label)
        self.setLayout(layout)

        # Set up a QTimer to update the brake level periodically
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.updateBrakeLevel)
        self.timer.start(1000)  # Update every second

    def updateImage(self, level):
        # Map slider value to image path
        image_paths = [
            "Downloads/ACS/ACS-empty.png",  # Image for brake level 0
            "Downloads/ACS/ACS-06.png",  # Image for brake level 1
            "Downloads/ACS/ACS-07.png",  # Image for brake level 2
            "Downloads/ACS/ACS-08.png",  # Image for brake level 3
            "Downloads/ACS/ACS-09.png"   # Image for brake level 4
        ]
        
        pixmap = QPixmap(image_paths[level])
        self.label.setPixmap(pixmap.scaled(self.label.size(), Qt.KeepAspectRatio))

    def updateBrakeLevel(self):
        try:
            with open('current_warning.txt', 'r') as f:
                brake_level = int(f.read().strip())
        except FileNotFoundError:
            brake_level = 0
        except Exception as e:
            print(f"An error occurred while reading brake level from file: {e}")
            brake_level = 0

        self.updateImage(brake_level)

def main():
    app = QApplication(sys.argv)
    ex = BrakeLevelApp()
    ex.show()

    # Create a ZED camera object
    zed = sl.Camera()

    # Set configuration parameters
    input_type = sl.InputType()
    if len(sys.argv) >= 2:
        input_type.set_from_svo_file(sys.argv[1])
    init = sl.InitParameters(input_t=input_type)
    init.camera_resolution = sl.RESOLUTION.HD1080
    init.depth_mode = sl.DEPTH_MODE.PERFORMANCE
    init.coordinate_units = sl.UNIT.MILLIMETER

    # Open the camera
    err = zed.open(init)
    if err != sl.ERROR_CODE.SUCCESS:
        print(repr(err))
        zed.close()
        sys.exit(1)

    runtime = sl.RuntimeParameters()

    image_size = zed.get_camera_information().camera_configuration.resolution
    image_size.width = image_size.width / 2
    image_size.height = image_size.height / 2

    image_zed = sl.Mat(image_size.width, image_size.height, sl.MAT_TYPE.U8_C4)
    depth_image_zed = sl.Mat(image_size.width, image_size.height, sl.MAT_TYPE.U8_C4)
    point_cloud = sl.Mat()

    key = ' '
    while key != 113:  # 'q' key to quit
        err = zed.grab(runtime)
        if err == sl.ERROR_CODE.SUCCESS:
            zed.retrieve_image(image_zed, sl.VIEW.RIGHT, sl.MEM.CPU, image_size)
            zed.retrieve_image(depth_image_zed, sl.VIEW.DEPTH, sl.MEM.CPU, image_size)
            zed.retrieve_measure(point_cloud, sl.MEASURE.XYZRGBA, sl.MEM.CPU, image_size)

            image_ocv = image_zed.get_data()
            depth_image_ocv = depth_image_zed.get_data()

            img = cv2.cvtColor(image_ocv, cv2.COLOR_BGRA2RGB)
            results = model(img, stream=True, classes=[0])

            nearest_distance = float('inf')
            nearest_person = None

            if results:
                for r in results:
                    boxes = r.boxes

                    for box in boxes:
                        x1, y1, x2, y2 = box.xyxy[0]
                        x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                        tengah_x = (x1 + x2) / 2
                        tengah_y = (y1 + y2) / 2
                        rounded_x = int(round(tengah_x))
                        rounded_y = int(round(tengah_y))
                        err, point_cloud_value = point_cloud.get_value(rounded_x, rounded_y)
                        distance = math.sqrt(point_cloud_value[0] ** 2 + point_cloud_value[1] ** 2 + point_cloud_value[2] ** 2)

                        if distance < nearest_distance:
                            nearest_distance = distance
                            nearest_person = box

            if nearest_person:
                x1, y1, x2, y2 = nearest_person.xyxy[0]
                x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                center_coordinates = (int((x1 + x2) / 2), int((y1 + y2) / 2))
                cv2.circle(image_ocv, center_coordinates, 10, (0, 255, 0), 3)
                cv2.rectangle(image_ocv, (x1, y1), (x2, y2), (0, 255, 0), 3)
                org = [x1, y1]
                font = cv2.FONT_HERSHEY_SIMPLEX
                fontScale = 1
                color = (0, 255, 0)
                thickness = 2
                cv2.putText(image_ocv, str(round(nearest_distance / 1000, 2)) + 'm', org, font, fontScale, color, thickness)

                nearest_distance_m = nearest_distance / 1000
                warning = distance_level(nearest_distance_m)
                pre_brake_percentage = start_warning_sequence(warning)

                log_to_excel(nearest_distance_m, warning, pre_brake_percentage)
                log_to_warning_file(warning)
            else:
                # If no person is detected, set warning level to 0 (default)
                warning = 0
                log_to_warning_file(warning)

            cv2.imshow("Image", image_ocv)
            cv2.imshow("Depth", depth_image_ocv)

            key = cv2.waitKey(1)

    zed.close()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
